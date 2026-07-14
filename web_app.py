#!/usr/bin/env python3
"""Local teacher-friendly web UI for the objective grading assistant."""

import csv
from dataclasses import dataclass
from email.parser import BytesParser
from email.policy import default as email_policy
import json
import mimetypes
import re
import sys
import tempfile
import traceback
import uuid
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Dict, List, Optional, Union
from urllib.parse import parse_qs, unquote, urlparse

PROJECT_ROOT = Path(__file__).resolve().parent
WEB_ROOT = PROJECT_ROOT / "web"
DATA_ROOT = PROJECT_ROOT / "data"
UPLOAD_ROOT = DATA_ROOT / "uploads"
CAPTURE_ROOT = DATA_ROOT / "captures"
REPORTS_ROOT = DATA_ROOT / "reports"
EXAMS_ROOT = DATA_ROOT / "exams"

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import roster_manager
from app.capture.mobile_web_camera_source import MAX_MOBILE_CAPTURE_REQUEST_BYTES
from app.application.product_preview_adapter import build_preview_data
from app.data_io import draft_to_review_rows, parse_answer_source, review_rows_to_answer_key_csv
from app.validators import has_blocking_errors
from app.workflow import make_run_id
from app.web_product import ProductWebController, UploadedFile


_PRODUCT_CONTROLLER: Optional[ProductWebController] = None


def product_controller() -> ProductWebController:
    global _PRODUCT_CONTROLLER
    if _PRODUCT_CONTROLLER is None:
        _PRODUCT_CONTROLLER = ProductWebController(DATA_ROOT / "local_app")
    return _PRODUCT_CONTROLLER


def json_bytes(data: object) -> bytes:
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")


def read_csv_rows(path: Path, limit: int = 8) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    return rows[:limit]


def table_file_to_csv(path: Path) -> Path:
    if path.suffix.lower() == ".csv":
        return path
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        return path
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("这台电脑暂时无法读取 Excel 文件，请另存为 CSV 后再上传。") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError("这个 Excel 文件暂时无法读取，请确认文件没有损坏，或另存为 CSV 后再上传。") from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    csv_path = path.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        for row in rows:
            writer.writerow(["" if value is None else value for value in row])
    return csv_path


def write_json(path: Path, data: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_json(path: Path, default: Optional[Dict[str, object]] = None) -> Dict[str, object]:
    if not path.exists():
        return default or {}
    return json.loads(path.read_text(encoding="utf-8"))


@dataclass
class FormField:
    value: str = ""
    filename: str = ""
    content: bytes = b""
    content_type: str = ""


FormData = Dict[str, Union[FormField, List[FormField]]]


def field_text(form: FormData, name: str, default: str = "") -> str:
    item = form.get(name)
    if isinstance(item, list):
        item = item[0] if item else None
    if item is None or item.filename:
        return default
    value = item.value
    return str(value or default).strip()


def save_upload(form: FormData, name: str, directory: Path, fallback: str = "") -> str:
    item = form.get(name)
    if item is None:
        return ""
    if isinstance(item, list):
        item = item[0]
    filename = Path(item.filename or fallback).name
    if not filename:
        return ""
    directory.mkdir(parents=True, exist_ok=True)
    target = directory / filename
    target.write_bytes(item.content)
    return filename


def load_roster_rows(class_name: str) -> List[Dict[str, str]]:
    roster = roster_manager.load_roster(class_name)
    return [{"student_id": student_id, "name": name} for student_id, name in sorted(roster.items())]


def build_unmatched_for_plain_submissions(class_name: str, submissions_path: Path) -> List[Dict[str, object]]:
    try:
        roster = roster_manager.load_roster(class_name)
    except Exception:
        roster = {}
    if not roster or not submissions_path.exists():
        return []
    rows = read_csv_rows(submissions_path, limit=100000)
    unmatched = []
    for index, row in enumerate(rows, start=2):
        student_id = row.get("student_id") or row.get("id") or row.get("学号") or row.get("考号") or ""
        student_id = str(student_id).strip()
        if student_id and student_id not in roster:
            unmatched.append(
                {
                    "recognized_student_id": student_id,
                    "row_number": index,
                    "message": "名单中没有找到这个学号",
                    "suggested_student_id": "",
                    "suggested_name": "",
                    "confidence": 0,
                    "action": "pending",
                }
            )
    return unmatched


def preview_session(session_dir: Path) -> Dict[str, object]:
    metadata = read_json(session_dir / "session.json")
    answer_key_path = session_dir / str(metadata["answer_key"])
    submissions_path = session_dir / str(metadata["submissions"])
    question_bank_name = str(metadata.get("question_bank") or "")
    question_bank_path = session_dir / question_bank_name if question_bank_name else None
    core_preview = build_preview_data(
        answer_key_path,
        submissions_path,
        question_bank_path,
    )
    validation_rows = list(core_preview["validation_rows"])
    unmatched = build_unmatched_for_plain_submissions(str(metadata.get("class_name", "")), submissions_path)
    for row in unmatched:
        validation_rows.append({"severity": "warning", "scope": "student_match", "item": row["recognized_student_id"], "message": row["message"]})
    return {
        "session_id": session_dir.name,
        "question_count": core_preview["question_count"],
        "student_count": core_preview["student_count"],
        "preview_rows": read_csv_rows(submissions_path, limit=5),
        "validation_rows": validation_rows,
        "blocking": has_blocking_errors(validation_rows),
        "unmatched_students": unmatched,
    }


def history_items() -> List[Dict[str, object]]:
    items: List[Dict[str, object]] = []
    for metadata_path in EXAMS_ROOT.rglob("exam_metadata.json"):
        directory = metadata_path.parent
        metadata = read_json(metadata_path)
        summary_path = directory / "summary.csv"
        student_count = metadata.get("student_count") or metadata.get("submission_count") or ""
        average = ""
        if summary_path.exists():
            rows = read_csv_rows(summary_path, limit=100000)
            scores = []
            for row in rows:
                try:
                    scores.append(float(row.get("score", 0)))
                except (TypeError, ValueError):
                    pass
            if scores:
                average = round(sum(scores) / len(scores), 2)
                student_count = len(scores)
        items.append(
            {
                "exam_name": metadata.get("exam_name", directory.name),
                "class_name": metadata.get("class_name", ""),
                "exam_date": metadata.get("exam_date", ""),
                "created_at": metadata.get("created_at", ""),
                "run_id": metadata.get("run_id", ""),
                "student_count": student_count,
                "average": average,
                "status": "已完成" if (directory / "index.html").exists() else "需检查",
                "index_url": f"/file?p={directory.relative_to(PROJECT_ROOT).as_posix()}/index.html" if (directory / "index.html").exists() else "",
                "dashboard_url": f"/file?p={directory.relative_to(PROJECT_ROOT).as_posix()}/advanced_dashboard.html" if (directory / "advanced_dashboard.html").exists() else "",
                "teaching_url": f"/file?p={directory.relative_to(PROJECT_ROOT).as_posix()}/teaching_plan.html" if (directory / "teaching_plan.html").exists() else "",
                "directory": str(directory),
            }
        )
    return sorted(items, key=lambda item: str(item.get("created_at") or item.get("exam_date")), reverse=True)


class WebHandler(BaseHTTPRequestHandler):
    server_version = "ObjectiveGraderWeb/1.0"

    def log_message(self, format: str, *args: object) -> None:
        return

    def send_bytes(self, body: bytes, content_type: str = "application/octet-stream", status: int = 200) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_json(self, data: object, status: int = 200) -> None:
        self.send_bytes(json_bytes(data), "application/json; charset=utf-8", status)

    def send_error_json(self, message: str, status: int = 400) -> None:
        self.send_json({"ok": False, "message": message}, status)

    def send_product_response(self, response) -> None:
        self.send_response(response.status)
        self.send_header("Content-Type", response.content_type)
        for name, value in response.headers.items():
            self.send_header(name, value)
        self.send_header("Content-Length", str(len(response.body)))
        self.end_headers()
        self.wfile.write(response.body)

    @staticmethod
    def product_form(form: FormData):
        fields = {}
        files = {}
        for name, item in form.items():
            if isinstance(item, list):
                item = item[0] if item else FormField()
            if item.filename:
                files[name] = UploadedFile(
                    item.filename,
                    item.content,
                    item.content_type,
                )
            else:
                fields[name] = item.value.strip()
        return fields, files

    def parse_form(self) -> FormData:
        content_type = self.headers.get("Content-Type", "")
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        form: FormData = {}
        if "multipart/form-data" not in content_type:
            for key, values in parse_qs(body.decode("utf-8", errors="replace")).items():
                form[key] = FormField(value=values[0] if values else "")
            return form

        raw = (
            f"Content-Type: {content_type}\r\n"
            f"MIME-Version: 1.0\r\n\r\n"
        ).encode("utf-8") + body
        message = BytesParser(policy=email_policy).parsebytes(raw)
        for part in message.iter_parts():
            disposition = part.get("Content-Disposition", "")
            if "form-data" not in disposition:
                continue
            name = part.get_param("name", header="content-disposition")
            if not name:
                continue
            filename = part.get_filename() or ""
            payload = part.get_payload(decode=True) or b""
            field = FormField(
                value=payload.decode(part.get_content_charset() or "utf-8", errors="replace") if not filename else "",
                filename=filename,
                content=payload,
                content_type=part.get_content_type(),
            )
            old = form.get(name)
            if old is None:
                form[name] = field
            elif isinstance(old, list):
                old.append(field)
            else:
                form[name] = [old, field]
        return form

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            product_response = product_controller().get(parsed.path)
            if product_response is not None:
                return self.send_product_response(product_response)
            if parsed.path == "/":
                return self.serve_static(WEB_ROOT / "index.html")
            if parsed.path.startswith("/static/"):
                return self.serve_static(WEB_ROOT / parsed.path.lstrip("/"))
            if parsed.path == "/file":
                return self.serve_project_file(parse_qs(parsed.query).get("p", [""])[0])
            if parsed.path == "/api/classes":
                return self.send_json({"ok": True, "classes": roster_manager.list_classes()})
            if parsed.path == "/api/class":
                class_name = parse_qs(parsed.query).get("class_name", [""])[0]
                return self.send_json({"ok": True, "students": load_roster_rows(class_name)})
            if parsed.path == "/api/classes/template":
                body = "student_id,name\n202601,张三\n202602,李四\n".encode("utf-8-sig")
                return self.send_bytes(body, "text/csv; charset=utf-8")
            if parsed.path == "/api/templates/answer":
                body = "question,answer,points,partial_credit,tags,difficulty\n1,A,1,false,集合,1\n2,BD,2,true,函数,3\n".encode("utf-8-sig")
                return self.send_bytes(body, "text/csv; charset=utf-8")
            if parsed.path == "/api/templates/submissions":
                body = "student_id,name,Q1,Q2\n202601,张三,A,BD\n202602,李四,B,D\n".encode("utf-8-sig")
                return self.send_bytes(body, "text/csv; charset=utf-8")
            if parsed.path == "/api/exams/history":
                return self.send_json({"ok": True, "items": history_items()})
            return self.send_error_json("没有找到这个页面。", 404)
        except Exception as exc:
            traceback.print_exc()
            return self.send_error_json(str(exc), 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if re.fullmatch(r"/sessions/[^/]+/capture/mobile-web", parsed.path):
                try:
                    request_length = int(self.headers.get("Content-Length", "0"))
                except ValueError:
                    return self.send_error_json("Content-Length 无效。", 400)
                if request_length > MAX_MOBILE_CAPTURE_REQUEST_BYTES:
                    return self.send_json(
                        {
                            "ok": False,
                            "message": "上传请求超过允许大小。",
                            "error_code": "REQUEST_TOO_LARGE",
                        },
                        413,
                    )
            if product_controller().handles(parsed.path):
                fields, files = self.product_form(self.parse_form())
                product_response = product_controller().post(
                    parsed.path,
                    fields,
                    files,
                )
                if product_response is not None:
                    return self.send_product_response(product_response)
            if parsed.path == "/api/classes/import":
                return self.handle_import_class()
            if parsed.path == "/api/exams/preview":
                return self.handle_preview()
            if parsed.path == "/api/exams/grade":
                return self.send_error_json(
                    "旧发布接口已停用，请通过考试会话的复核与最终发布页面操作。",
                    410,
                )
            if parsed.path == "/api/answer/parse":
                return self.handle_answer_parse()
            if parsed.path == "/api/answer/confirm":
                return self.handle_answer_confirm()
            if parsed.path == "/api/captures/upload":
                return self.handle_capture_upload()
            return self.send_error_json("没有找到这个操作。", 404)
        except Exception as exc:
            traceback.print_exc()
            return self.send_error_json(str(exc), 500)

    def serve_static(self, path: Path) -> None:
        path = path.resolve()
        if not str(path).startswith(str(WEB_ROOT.resolve())) or not path.exists() or not path.is_file():
            return self.send_error_json("没有找到这个文件。", 404)
        content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        if path.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif path.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif path.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        self.send_bytes(path.read_bytes(), content_type)

    def serve_project_file(self, relative_path: str) -> None:
        relative_path = unquote(relative_path or "")
        path = (PROJECT_ROOT / relative_path).resolve()
        if not str(path).startswith(str(PROJECT_ROOT.resolve())) or not path.exists() or not path.is_file():
            return self.send_error_json("没有找到报告文件。", 404)
        content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        if path.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        self.send_bytes(path.read_bytes(), content_type)

    def handle_import_class(self) -> None:
        form = self.parse_form()
        class_name = field_text(form, "class_name")
        if not class_name:
            return self.send_error_json("请填写班级名称。")
        upload_dir = Path(tempfile.mkdtemp(prefix="roster_", dir=str(UPLOAD_ROOT)))
        filename = save_upload(form, "roster", upload_dir, "roster.csv")
        if not filename:
            return self.send_error_json("请上传学生名单文件。")
        result = roster_manager.import_roster(upload_dir / filename, class_name)
        self.send_json({"ok": True, "message": "名单导入完成。", "result": result, "classes": roster_manager.list_classes()})

    def handle_preview(self) -> None:
        form = self.parse_form()
        session_id = field_text(form, "session_id") or uuid.uuid4().hex[:12]
        session_dir = UPLOAD_ROOT / session_id
        session_dir.mkdir(parents=True, exist_ok=True)
        existing = read_json(session_dir / "session.json")
        answer_key = save_upload(form, "answer_key", session_dir, "answer_key.csv") or str(existing.get("answer_key") or "")
        submissions = save_upload(form, "submissions", session_dir, "submissions.csv")
        question_bank = save_upload(form, "question_bank", session_dir, "question_bank.csv")
        if submissions:
            submissions = table_file_to_csv(session_dir / submissions).name
        if question_bank:
            question_bank = table_file_to_csv(session_dir / question_bank).name
        if not answer_key or not submissions:
            return self.send_error_json("请先确认标准答案，并上传学生作答文件。")
        metadata = {
            "class_name": field_text(form, "class_name") or str(existing.get("class_name") or ""),
            "exam_name": field_text(form, "exam_name") or str(existing.get("exam_name") or ""),
            "exam_date": field_text(form, "exam_date") or str(existing.get("exam_date") or ""),
            "subject": field_text(form, "subject") or str(existing.get("subject") or ""),
            "answer_key": answer_key,
            "submissions": submissions,
            "question_bank": question_bank or str(existing.get("question_bank") or ""),
            "created_at": make_run_id(),
        }
        write_json(session_dir / "session.json", metadata)
        preview = preview_session(session_dir)
        preview["ok"] = True
        self.send_json(preview)

    def handle_answer_parse(self) -> None:
        form = self.parse_form()
        session_id = field_text(form, "session_id") or uuid.uuid4().hex[:12]
        session_dir = UPLOAD_ROOT / session_id
        session_dir.mkdir(parents=True, exist_ok=True)
        source_name = save_upload(form, "answer_source", session_dir, "answer_source")
        manual_text = field_text(form, "manual_text")
        if not source_name and not manual_text:
            return self.send_error_json("请上传标准答案文件，或粘贴/填写标准答案。")
        if manual_text:
            source_path = session_dir / "manual_answer.txt"
            source_path.write_text(manual_text, encoding="utf-8-sig")
        else:
            source_path = session_dir / source_name
        draft = parse_answer_source(source_path)
        write_json(session_dir / "answer_draft.json", draft)
        metadata = read_json(session_dir / "session.json")
        for key in ["class_name", "exam_name", "exam_date", "subject"]:
            value = field_text(form, key)
            if value:
                metadata[key] = value
        metadata["answer_source"] = str(source_path)
        write_json(session_dir / "session.json", metadata)
        self.send_json({"ok": True, "session_id": session_id, "draft": draft, "review_rows": draft_to_review_rows(draft)})

    def handle_answer_confirm(self) -> None:
        length = int(self.headers.get("Content-Length", "0"))
        payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
        session_id = str(payload.get("session_id", ""))
        rows = payload.get("rows", [])
        if not session_id or not isinstance(rows, list):
            return self.send_error_json("请先生成标准答案草稿。")
        session_dir = (UPLOAD_ROOT / session_id).resolve()
        if not str(session_dir).startswith(str(UPLOAD_ROOT.resolve())):
            return self.send_error_json("没有找到这次考试的临时文件。")
        answer_key_path = session_dir / "confirmed_answer_key.csv"
        review_rows_to_answer_key_csv(rows, answer_key_path)
        metadata = read_json(session_dir / "session.json")
        metadata["answer_key"] = "confirmed_answer_key.csv"
        metadata["answer_confirmed"] = True
        write_json(session_dir / "session.json", metadata)
        self.send_json({"ok": True, "message": "标准答案已确认，可以导入学生作答。", "session_id": session_id})

    def handle_capture_upload(self) -> None:
        length = int(self.headers.get("Content-Length", "0"))
        payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
        session_id = str(payload.get("session_id") or uuid.uuid4().hex[:12])
        images = payload.get("images", [])
        if not isinstance(images, list) or not images:
            return self.send_error_json("还没有拍摄图片。")
        import base64

        capture_dir = CAPTURE_ROOT / session_id
        capture_dir.mkdir(parents=True, exist_ok=True)
        manifest = {"images": []}
        for index, image in enumerate(images, start=1):
            if not isinstance(image, dict):
                continue
            data_url = str(image.get("dataUrl") or "")
            if "," not in data_url:
                continue
            filename = f"answer_sheet_{index:03d}.jpg"
            raw = base64.b64decode(data_url.split(",", 1)[1])
            (capture_dir / filename).write_bytes(raw)
            manifest["images"].append(
                {
                    "filename": filename,
                    "sourceMode": image.get("sourceMode", ""),
                    "deviceLabel": image.get("deviceLabel", ""),
                    "capturedAt": image.get("capturedAt", ""),
                    "status": "pending",
                }
            )
        write_json(capture_dir / "capture_manifest.json", manifest)
        self.send_json({"ok": True, "session_id": session_id, "saved_count": len(manifest["images"]), "message": "拍照图片已保存，识别前请先确认。"})

def run(host: str = "127.0.0.1", port: int = 8765) -> None:
    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    CAPTURE_ROOT.mkdir(parents=True, exist_ok=True)
    REPORTS_ROOT.mkdir(parents=True, exist_ok=True)
    EXAMS_ROOT.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((host, port), WebHandler)
    print(f"客观题批改助手已启动：http://{host}:{port}")
    print("按 Ctrl+C 结束服务。")
    server.serve_forever()


if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    run(port=port)
