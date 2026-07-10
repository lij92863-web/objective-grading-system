import csv
import uuid
from pathlib import Path

from app.classes.class_repository import ClassRepository
from app.storage.local_db import LocalDatabase
from app.storage.migrations import initialize_database, utc_now
from app.storage.transaction import transaction

from .roster_mapping import RosterColumnMapping, detect_mapping
from .roster_repository import RosterRepository
from .roster_validator import (
    RosterImportResult,
    RosterImportState,
    Student,
    validate_rows,
)


class RosterImporter:
    def __init__(self, database: LocalDatabase) -> None:
        self.database = database
        self.classes = ClassRepository()
        self.rosters = RosterRepository()
        initialize_database(database)

    def import_file(
        self,
        class_id: str,
        path: Path,
        mapping: RosterColumnMapping | None = None,
    ) -> RosterImportResult:
        headers, rows = self._read(Path(path))
        selected = mapping or detect_mapping(headers)
        if selected is None:
            return RosterImportResult(
                RosterImportState.COLUMN_MAPPING_REQUIRED,
                0,
                (),
                (),
                tuple(headers),
            )
        normalized, warnings, blocking = validate_rows(rows, selected)
        if blocking:
            return RosterImportResult(
                RosterImportState.BLOCKED,
                0,
                tuple(warnings),
                tuple(blocking),
                tuple(headers),
                selected,
            )
        now = utc_now()
        students = [
            Student(uuid.uuid4().hex, class_id, number, name, (), "ACTIVE", now, now)
            for number, name in normalized
        ]
        with transaction(self.database) as connection:
            if self.classes.get(connection, class_id) is None:
                raise ValueError("class does not exist")
            self.rosters.add_many(connection, students)
        return RosterImportResult(
            RosterImportState.IMPORTED,
            len(students),
            tuple(warnings),
            (),
            tuple(headers),
            selected,
        )

    def list_students(self, class_id: str) -> list[Student]:
        with self.database.connection() as connection:
            return self.rosters.list_for_class(connection, class_id)

    @staticmethod
    def _read(path: Path) -> tuple[list[str], list[dict[str, object]]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                if not reader.fieldnames:
                    raise ValueError("roster has no header row")
                return list(reader.fieldnames), list(reader)
        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("Excel support is unavailable; save as CSV") from exc
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                values = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
            if not values:
                raise ValueError("roster is empty")
            headers = [str(value or "").strip() for value in values[0]]
            rows = [dict(zip(headers, row)) for row in values[1:]]
            return headers, rows
        raise ValueError("roster must be CSV or XLSX")
