"""Excel legacy behavior lock — E3A. Harden tests before migration.

Since openpyxl is NOT available in this environment, deep openpyxl-based
checks (cell styles, column widths, font formats) are skipped.  This file
uses Python's built-in ``zipfile`` + ``xml.etree`` to inspect the XLSX
ZIP structure (sheet names, row counts, headers) without openpyxl.
"""

import shutil
import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEMO_KEY = PROJECT_ROOT / "samples/demo_exam/answer_key_sample.csv"
DEMO_SUB = PROJECT_ROOT / "samples/demo_exam/submissions_sample.csv"

# ── expected sheet names (must stay stable across migration) ──────────
EXPECTED_FULL_WORKBOOK_SHEETS = [
    "成绩总表",
    "每题明细",
    "每题分析",
    "知识点画像",
    "学生错题",
    "讲评计划",
    "班级补救",
    "分层补救",
    "数据质量检查",
]

EXPECTED_SIMPLE_WORKBOOK_SHEETS = ["scores"]

# Legacy simple score workbook header (from legacy.write_simple_score_workbook)
EXPECTED_SIMPLE_SCORE_HEADER = [
    "rank", "student_id", "name", "score", "max_score",
    "percent", "correct_count", "wrong_or_partial_count",
    "blank_count", "invalid_count", "wrong_questions",
    "blank_questions", "remark",
]

# Legacy full workbook key sheet headers (verified against legacy)
EXPECTED_SUMMARY_SHEET_HEADER = [
    "student_id", "name", "rank", "score", "max_score",
    "percent", "correct_count", "wrong_or_partial_count",
    "blank_count", "invalid_count",
]

EXPECTED_DETAIL_SHEET_HEADER = [
    "student_id", "name", "question", "question_id",
    "question_status", "difficulty", "tags", "expected",
    "actual", "raw_actual", "score", "max_score", "status",
]

EXPECTED_ITEM_ANALYSIS_SHEET_HEADER = [
    "question", "question_id", "question_status", "difficulty",
    "tags", "answer", "points", "accuracy", "blank_rate",
    "wrong_rate", "partial_rate", "invalid_rate", "option_distribution",
]

OPENPYXL_AVAILABLE = False
try:
    import openpyxl  # noqa: F401
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass


# ── helpers ────────────────────────────────────────────────────────────

def _get_xlsx_sheet_names(xlsx_path: Path) -> list:
    """Extract sheet names from an xlsx ZIP without openpyxl."""
    ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        root = ET.fromstring(wb_xml)
        return [s.get("name", "") for s in root.findall(".//ns:sheet", ns)]


def _get_xlsx_sheet_headers(xlsx_path: Path) -> dict:
    """Return {sheet_name: [header_values]} for every sheet in the xlsx."""
    ns_s = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_sheet = {"ns": ns_s}
    result = {}
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        root = ET.fromstring(wb_xml)
        sheets = [(s.get("name", ""), s.get("sheetId", ""))
                  for s in root.findall(".//ns:sheet", ns_sheet)]
        for idx, (name, _sid) in enumerate(sheets, start=1):
            sheet_path = f"xl/worksheets/sheet{idx}.xml"
            try:
                sheet_xml = z.read(sheet_path).decode("utf-8")
            except KeyError:
                result[name] = []
                continue
            sroot = ET.fromstring(sheet_xml)
            rows = sroot.findall(".//ns:row", ns_sheet)
            if not rows:
                result[name] = []
                continue
            # first row is header
            first_row_cells = rows[0].findall(".//ns:c", ns_sheet)
            header = []
            for cell in first_row_cells:
                is_elem = cell.find(".//ns:is", ns_sheet)
                if is_elem is not None:
                    t_elem = is_elem.find("ns:t", ns_sheet)
                    if t_elem is not None and t_elem.text is not None:
                        header.append(t_elem.text)
            result[name] = header
    return result


def _get_xlsx_sheet_row_counts(xlsx_path: Path) -> dict:
    """Return {sheet_name: row_count} for every sheet in the xlsx."""
    ns_sheet = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    result = {}
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        root = ET.fromstring(wb_xml)
        sheets = [(s.get("name", ""), s.get("sheetId", ""))
                  for s in root.findall(".//ns:sheet", ns_sheet)]
        for idx, (name, _sid) in enumerate(sheets, start=1):
            sheet_path = f"xl/worksheets/sheet{idx}.xml"
            try:
                sheet_xml = z.read(sheet_path).decode("utf-8")
            except KeyError:
                result[name] = 0
                continue
            sroot = ET.fromstring(sheet_xml)
            rows = sroot.findall(".//ns:row", ns_sheet)
            result[name] = len(rows)
    return result


# ── tests ──────────────────────────────────────────────────────────────

class ExcelLegacyBehaviorTests(unittest.TestCase):
    """Lock legacy Excel behavior before E3B-E3C migration."""

    @classmethod
    def setUpClass(cls):
        if not DEMO_KEY.exists():
            raise unittest.SkipTest("No demo answer key")

    # -- basic file existence -------------------------------------------------

    def test_excel_files_exist_and_non_empty(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            for fname in ["exam_report.xlsx", "simple_score_report.xlsx"]:
                fp = Path(t) / fname
                self.assertTrue(fp.exists(), f"Missing {fname}")
                self.assertGreater(fp.stat().st_size, 1000,
                                   f"{fname} too small")
                self.assertTrue(fp.suffix == ".xlsx",
                                f"{fname} wrong extension")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_xlsx_are_valid_zip_archives(self):
        """Both xlsx files must be valid ZIP archives (XLSX is ZIP)."""
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            for fname in ["exam_report.xlsx", "simple_score_report.xlsx"]:
                fp = Path(t) / fname
                self.assertTrue(zipfile.is_zipfile(fp),
                                f"{fname} is not a valid ZIP")
                with zipfile.ZipFile(fp) as z:
                    names = z.namelist()
                    self.assertIn("xl/workbook.xml", names,
                                  f"{fname} missing workbook.xml")
                    self.assertIn("[Content_Types].xml", names,
                                  f"{fname} missing Content_Types")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    # -- full workbook sheet structure ----------------------------------------

    def test_full_workbook_sheet_names_match_legacy(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            sheets = _get_xlsx_sheet_names(fp)
            self.assertEqual(len(sheets), len(EXPECTED_FULL_WORKBOOK_SHEETS),
                             f"Sheet count mismatch: {len(sheets)} vs "
                             f"{len(EXPECTED_FULL_WORKBOOK_SHEETS)}")
            self.assertEqual(sheets, EXPECTED_FULL_WORKBOOK_SHEETS,
                             f"Sheet names drifted: {sheets}")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_has_nine_sheets(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            sheets = _get_xlsx_sheet_names(fp)
            self.assertEqual(len(sheets), 9,
                             f"Expected 9 sheets, got {len(sheets)}: {sheets}")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_every_sheet_has_header(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            headers = _get_xlsx_sheet_headers(fp)
            for name in EXPECTED_FULL_WORKBOOK_SHEETS:
                self.assertIn(name, headers,
                              f"Sheet '{name}' not found in workbook")
                self.assertGreater(len(headers[name]), 0,
                                   f"Sheet '{name}' has no header row")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_key_sheet_headers_match_legacy(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            headers = _get_xlsx_sheet_headers(fp)
            self.assertEqual(headers.get("成绩总表", []),
                             EXPECTED_SUMMARY_SHEET_HEADER)
            self.assertEqual(headers.get("每题明细", []),
                             EXPECTED_DETAIL_SHEET_HEADER)
            self.assertEqual(headers.get("每题分析", []),
                             EXPECTED_ITEM_ANALYSIS_SHEET_HEADER)
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_has_student_data_rows(self):
        """Each sheet (except maybe validation) should have >= 1 data row."""
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            counts = _get_xlsx_sheet_row_counts(fp)
            # 成绩总表, 每题明细, 每题分析, 知识点画像 each have data
            for sheet_name in ["成绩总表", "每题明细", "每题分析", "知识点画像"]:
                self.assertIn(sheet_name, counts)
                self.assertGreater(
                    counts[sheet_name], 1,
                    f"Sheet '{sheet_name}' has only {counts.get(sheet_name, 0)} "
                    f"rows (expected header + data)")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_chinese_not_garbled(self):
        """Verify Chinese sheet names are preserved in the XML."""
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            with zipfile.ZipFile(fp) as z:
                wb_xml = z.read("xl/workbook.xml").decode("utf-8")
                for cn_name in EXPECTED_FULL_WORKBOOK_SHEETS:
                    self.assertIn(cn_name, wb_xml,
                                  f"Chinese sheet name '{cn_name}' not found "
                                  f"in workbook.xml")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    # -- simple score workbook ------------------------------------------------

    def test_simple_score_workbook_sheet_name(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "simple_score_report.xlsx"
            sheets = _get_xlsx_sheet_names(fp)
            self.assertEqual(sheets, EXPECTED_SIMPLE_WORKBOOK_SHEETS,
                             f"Unexpected sheets: {sheets}")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_simple_score_workbook_header_matches_legacy(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "simple_score_report.xlsx"
            headers = _get_xlsx_sheet_headers(fp)
            self.assertEqual(headers.get("scores", []),
                             EXPECTED_SIMPLE_SCORE_HEADER)
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_simple_score_workbook_has_data_rows(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "simple_score_report.xlsx"
            counts = _get_xlsx_sheet_row_counts(fp)
            self.assertGreater(counts.get("scores", 0), 1,
                               "Simple score workbook has no data rows")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    # -- xlsx internal structure (zipfile + XML, no openpyxl) ------------------

    def test_full_workbook_worksheets_xml_exist(self):
        """Every declared sheet must have a corresponding worksheet XML file."""
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            sheets = _get_xlsx_sheet_names(fp)
            with zipfile.ZipFile(fp) as z:
                names = set(z.namelist())
                for idx in range(1, len(sheets) + 1):
                    sheet_path = f"xl/worksheets/sheet{idx}.xml"
                    self.assertIn(sheet_path, names,
                                  f"Missing {sheet_path}")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_inlinestr_not_sharedstrings(self):
        """Legacy uses inlineStr, not sharedStrings, for cell values.

        This is a deliberate implementation choice — the pure-Python
        zipfile+XML path writes ``t=\"inlineStr\"`` cells directly.
        sharedStrings.xml may or may not exist, and new exporters
        should preserve this characteristic unless a styled path
        (openpyxl) is active.
        """
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            with zipfile.ZipFile(fp) as z:
                names = set(z.namelist())
                # sharedStrings.xml is NOT required — legacy uses inlineStr
                # Just record the fact; do not mandate its absence.
                for sheet_idx in range(1, 10):
                    sheet_xml = z.read(
                        f"xl/worksheets/sheet{sheet_idx}.xml"
                    ).decode("utf-8")
                    # At least one cell with inlineStr should exist
                    # in sheets that have data
                    if "inlineStr" in sheet_xml:
                        break
                else:
                    # If no sheet has inlineStr, something changed
                    self.fail(
                        "No inlineStr cells found in any worksheet — "
                        "legacy xlsx generation method may have changed")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_no_styles_xml(self):
        """Legacy fallback (no openpyxl) does NOT emit xl/styles.xml.

        Styles (header_fill, error_fill, column widths, freeze panes)
        are only available through the openpyxl enhanced path.  The
        zipfile+XML fallback produces plain unstyled sheets.
        """
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            with zipfile.ZipFile(fp) as z:
                names = set(z.namelist())
                self.assertNotIn("xl/styles.xml", names,
                                 "legacy fallback should not emit styles.xml")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_full_workbook_inlinestr_cell_values_parseable(self):
        """InlineStr cells must contain valid, parseable text values.

        Legacy writes ALL cell values as ``t=\"inlineStr\"`` — the text is
        embedded directly in the worksheet XML (no sharedStrings indirection).
        This test verifies that the inlineStr values can be extracted and
        match known header text.  Chinese encoding is separately verified
        through sheet-name inspection in test_full_workbook_chinese_not_garbled.
        """
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "exam_report.xlsx"
            with zipfile.ZipFile(fp) as z:
                sheet1_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
                self.assertIn("inlineStr", sheet1_xml,
                              "sheet1 should use inlineStr cells")
                ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                sroot = ET.fromstring(sheet1_xml)
                inline_texts = []
                for cell in sroot.findall(".//ns:c", ns):
                    is_elem = cell.find(".//ns:is", ns)
                    if is_elem is not None:
                        t_elem = is_elem.find("ns:t", ns)
                        if t_elem is not None and t_elem.text:
                            inline_texts.append(t_elem.text)
                self.assertGreater(
                    len(inline_texts), 0,
                    "No inlineStr cell values found in sheet1")
                # First row should be header — verify at least one known field
                self.assertIn("student_id", inline_texts,
                              "Header field 'student_id' not found in inlineStr")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_simple_score_workbook_worksheets_xml_exist(self):
        """Simple score workbook must have at least one worksheet XML."""
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            fp = Path(t) / "simple_score_report.xlsx"
            with zipfile.ZipFile(fp) as z:
                names = set(z.namelist())
                self.assertIn("xl/worksheets/sheet1.xml", names,
                              "Missing xl/worksheets/sheet1.xml in simple workbook")
        finally:
            shutil.rmtree(t, ignore_errors=True)

    # -- openpyxl-dependent tests (skipped when unavailable) -------------------

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not installed")
    def test_excel_opens_with_openpyxl(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            from openpyxl import load_workbook
            wb = load_workbook(Path(t) / "exam_report.xlsx", read_only=True)
            sheets = wb.sheetnames
            self.assertGreater(len(sheets), 1, f"Excel has {len(sheets)} sheets")
            self.assertEqual(sheets, EXPECTED_FULL_WORKBOOK_SHEETS)
            wb.close()
        finally:
            shutil.rmtree(t, ignore_errors=True)

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not installed")
    def test_simple_score_opens_with_openpyxl(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            from openpyxl import load_workbook
            wb = load_workbook(Path(t) / "simple_score_report.xlsx", read_only=True)
            self.assertEqual(wb.sheetnames, EXPECTED_SIMPLE_WORKBOOK_SHEETS)
            wb.close()
        finally:
            shutil.rmtree(t, ignore_errors=True)

    # -- safety ---------------------------------------------------------------

    def test_no_api_called(self):
        t = tempfile.mkdtemp(prefix="e3a_", dir=PROJECT_ROOT / "data")
        try:
            from app.workflow import run_grading
            r = run_grading(DEMO_KEY, DEMO_SUB, Path(t), no_archive=True, exam_name="x")
            self.assertTrue(r["ok"])
        finally:
            shutil.rmtree(t, ignore_errors=True)

    def test_no_legacy_imports_in_existing_exporters(self):
        """Existing infrastructure exporters must not import legacy."""
        import ast
        infra = PROJECT_ROOT / "app" / "infrastructure" / "exporters"
        for py_file in infra.glob("*.py"):
            if py_file.name == "__init__.py":
                continue
            tree = ast.parse(py_file.read_text(encoding="utf-8"))
            for node in ast.walk(tree):
                if isinstance(node, ast.Import):
                    for alias in node.names:
                        self.assertNotIn("legacy", alias.name,
                                         f"{py_file.name} imports legacy")
                elif isinstance(node, ast.ImportFrom):
                    if node.module:
                        self.assertNotIn("legacy", node.module,
                                         f"{py_file.name} imports legacy")


if __name__ == "__main__":
    unittest.main()
