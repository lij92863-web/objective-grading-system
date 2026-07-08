#!/usr/bin/env python3
"""Class roster management for the objective grader.

Teachers keep using readable Chinese class names. Files are stored under stable
ASCII class_id directories so copied or zipped project folders are less likely
to hit encoding problems on another computer.
"""

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


PROJECT_ROOT = Path(__file__).resolve().parent
DATA_ROOT = PROJECT_ROOT / "data"
CLASSES_ROOT = DATA_ROOT / "classes"
INDEX_PATH = CLASSES_ROOT / "classes_index.json"
DEFAULT_CLASS_PATH = CLASSES_ROOT / "default_class.json"

ID_HEADERS = {
    "student_id",
    "id",
    "学生编号",
    "学号",
    "编号",
    "考号",
    "准考证号",
    "座号",
    "座位号",
}
NAME_HEADERS = {"name", "姓名", "学生姓名", "名字"}
CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
INVALID_PATH_CHARS_RE = re.compile(r'[<>:"/\\|?*]')
CLASS_ID_RE = re.compile(r"^class_(\d{3,})$")


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def normalize_header(value: object) -> str:
    return str(value or "").replace("\u3000", " ").strip().lower()


def normalize_class_name(class_name: str) -> str:
    cleaned = (class_name or "").strip()
    if not cleaned:
        raise ValueError("班级名称不能为空。")
    if INVALID_PATH_CHARS_RE.search(cleaned):
        raise ValueError('班级名称不能包含这些符号：<>:"/\\|?*')
    return cleaned


def read_json(path: Path, default: Dict[str, object]) -> Dict[str, object]:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, data: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_dicts(path: Path, fieldnames: List[str], rows: List[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def _read_metadata(directory: Path) -> Dict[str, object]:
    return read_json(directory / "class_metadata.json", {}) if (directory / "class_metadata.json").exists() else {}


def _roster_count(directory: Path) -> int:
    roster_path = directory / "roster.csv"
    if not roster_path.exists():
        return 0
    with roster_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return sum(1 for _row in reader)


def _entry(
    class_id: str,
    class_name: str,
    student_count: int,
    updated_at: str,
) -> Dict[str, object]:
    return {
        "class_id": class_id,
        "class_name": class_name,
        "student_count": student_count,
        "updated_at": updated_at,
    }


def _upsert_entry(classes: List[Dict[str, object]], entry: Dict[str, object]) -> None:
    class_id = str(entry.get("class_id", ""))
    class_name = str(entry.get("class_name", ""))
    for index, item in enumerate(classes):
        if item.get("class_id") == class_id or item.get("class_name") == class_name:
            classes[index] = entry
            return
    classes.append(entry)


def next_class_id(index: Dict[str, object]) -> str:
    max_number = 0
    for item in index.get("classes", []):
        if not isinstance(item, dict):
            continue
        match = CLASS_ID_RE.match(str(item.get("class_id", "")))
        if match:
            max_number = max(max_number, int(match.group(1)))
    return f"class_{max_number + 1:03d}"


def _next_available_class_id(classes_root: Path, classes: List[Dict[str, object]]) -> str:
    index = {"classes": list(classes)}
    while True:
        class_id = next_class_id(index)
        if not (classes_root / class_id).exists():
            return class_id
        index["classes"].append({"class_id": class_id})


def _write_class_metadata(
    directory: Path,
    class_id: str,
    class_name: str,
    student_count: int,
    updated_at: str,
    source_file: str = "",
    notes: str = "",
) -> None:
    old = _read_metadata(directory)
    metadata = {
        "class_id": class_id,
        "class_name": class_name,
        "created_at": old.get("created_at", updated_at),
        "updated_at": updated_at,
        "student_count": student_count,
        "source_file": source_file or old.get("source_file", ""),
        "notes": notes if notes else old.get("notes", ""),
    }
    write_json(directory / "class_metadata.json", metadata)


def _migrate_legacy_class_dir(classes_root: Path, legacy_dir: Path, class_name: str, classes: List[Dict[str, object]]) -> Tuple[str, Path]:
    class_id = _next_available_class_id(classes_root, classes)
    target_dir = classes_root / class_id
    if not target_dir.exists():
        legacy_dir.rename(target_dir)
        return class_id, target_dir
    return legacy_dir.name, legacy_dir


def load_classes_index(classes_root: Path = CLASSES_ROOT) -> Dict[str, object]:
    classes_root.mkdir(parents=True, exist_ok=True)
    raw_index = read_json(classes_root / "classes_index.json", {"classes": []})
    classes: List[Dict[str, object]] = []
    timestamp = now_iso()

    for item in raw_index.get("classes", []):
        if not isinstance(item, dict):
            continue
        class_id = str(item.get("class_id", "")).strip()
        class_name = str(item.get("class_name", "")).strip()
        if class_id:
            directory = classes_root / class_id
            metadata = _read_metadata(directory)
            class_name = class_name or str(metadata.get("class_name", "")).strip() or class_id
            updated_at = str(metadata.get("updated_at") or item.get("updated_at") or timestamp)
            student_count = _roster_count(directory) or int(item.get("student_count") or 0)
            _upsert_entry(classes, _entry(class_id, class_name, student_count, updated_at))
            continue
        if class_name:
            legacy_dir = classes_root / class_name
            if legacy_dir.exists() and legacy_dir.is_dir():
                class_id, directory = _migrate_legacy_class_dir(classes_root, legacy_dir, class_name, classes)
                student_count = _roster_count(directory) or int(item.get("student_count") or 0)
                updated_at = str(item.get("updated_at") or timestamp)
                _write_class_metadata(directory, class_id, class_name, student_count, updated_at)
                _upsert_entry(classes, _entry(class_id, class_name, student_count, updated_at))

    known_ids = {str(item.get("class_id", "")) for item in classes}
    for directory in classes_root.iterdir():
        if not directory.is_dir() or directory.name == "__pycache__":
            continue
        if directory.name in known_ids:
            continue
        metadata = _read_metadata(directory)
        class_name = str(metadata.get("class_name") or directory.name).strip()
        if not class_name:
            continue
        if CLASS_ID_RE.match(directory.name):
            class_id = directory.name
        elif (directory / "roster.csv").exists():
            class_id, directory = _migrate_legacy_class_dir(classes_root, directory, class_name, classes)
        else:
            continue
        student_count = _roster_count(directory)
        updated_at = str(metadata.get("updated_at") or timestamp)
        _write_class_metadata(directory, class_id, class_name, student_count, updated_at)
        _upsert_entry(classes, _entry(class_id, class_name, student_count, updated_at))

    classes.sort(key=lambda item: str(item.get("class_id", "")))
    index = {"classes": classes}
    save_classes_index(classes_root, index)
    return index


def save_classes_index(classes_root: Path, index: Dict[str, object]) -> None:
    write_json(classes_root / "classes_index.json", index)


def ensure_classes_root(classes_root: Path = CLASSES_ROOT) -> None:
    classes_root.mkdir(parents=True, exist_ok=True)
    if not (classes_root / "classes_index.json").exists():
        write_json(classes_root / "classes_index.json", {"classes": []})
    if not (classes_root / "default_class.json").exists():
        write_json(classes_root / "default_class.json", {"default_class_id": "", "default_class_name": ""})
    load_classes_index(classes_root)


def find_class_entry(classes_root: Path, class_name: str) -> Dict[str, object]:
    query = normalize_class_name(class_name)
    index = load_classes_index(classes_root)
    for item in index.get("classes", []):
        if item.get("class_name") == query or item.get("class_id") == query:
            return dict(item)
    raise FileNotFoundError(f"未找到班级“{query}”，请先导入该班级学生名单。")


def class_dir(class_name: str, classes_root: Path = CLASSES_ROOT) -> Path:
    entry = find_class_entry(classes_root, class_name)
    return classes_root / str(entry["class_id"])


def _get_or_create_class_entry(class_name: str, classes_root: Path) -> Dict[str, object]:
    class_name = normalize_class_name(class_name)
    try:
        return find_class_entry(classes_root, class_name)
    except FileNotFoundError:
        index = load_classes_index(classes_root)
        classes = list(index.get("classes", []))
        class_id = _next_available_class_id(classes_root, classes)
        timestamp = now_iso()
        directory = classes_root / class_id
        directory.mkdir(parents=True, exist_ok=True)
        entry = _entry(class_id, class_name, 0, timestamp)
        _upsert_entry(classes, entry)
        save_classes_index(classes_root, {"classes": classes})
        _write_class_metadata(directory, class_id, class_name, 0, timestamp)
        return entry


def find_column(headers: Iterable[str], aliases: set) -> Optional[str]:
    for header in headers:
        if normalize_header(header) in aliases:
            return header
    return None


def has_scientific_notation(text: str) -> bool:
    return bool(re.fullmatch(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+", text.strip()))


def normalize_student_id(value: object) -> Tuple[str, List[str]]:
    warnings: List[str] = []
    if value is None:
        return "", warnings
    text = str(value).replace("\u3000", " ").strip()
    if not text:
        return "", warnings
    if CONTROL_CHAR_RE.search(text):
        warnings.append("学号中含有不可见字符，已自动清理")
        text = CONTROL_CHAR_RE.sub("", text)
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0], warnings
    if has_scientific_notation(text):
        try:
            restored = format(Decimal(text), "f")
            if "." in restored:
                restored = restored.rstrip("0").rstrip(".")
            warnings.append("学号像科学计数法，已尝试还原")
            return restored, warnings
        except InvalidOperation:
            warnings.append("学号像科学计数法，但无法自动还原")
    return text, warnings


def normalize_name(value: object) -> Tuple[str, List[str]]:
    warnings: List[str] = []
    text = str(value or "").replace("\u3000", " ").strip()
    if CONTROL_CHAR_RE.search(text):
        warnings.append("姓名中含有不可见字符，已自动清理")
        text = CONTROL_CHAR_RE.sub("", text)
    return text, warnings


def read_csv_roster_input(path: Path) -> Tuple[List[Dict[str, object]], List[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError(f"{path} 没有表头行。")
        return list(reader), list(reader.fieldnames)


def read_xlsx_roster_input(path: Path) -> Tuple[List[Dict[str, object]], List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 .xlsx 名单需要 openpyxl。请先安装：pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_values = next(rows_iter)
    except StopIteration as exc:
        raise ValueError(f"{path} 是空文件。") from exc
    headers = [str(value or "").strip() for value in header_values]
    rows: List[Dict[str, object]] = []
    for values in rows_iter:
        rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})
    workbook.close()
    return rows, headers


def read_roster_input(path: Path) -> Tuple[List[Dict[str, object]], List[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_roster_input(path)
    if suffix == ".xlsx":
        return read_xlsx_roster_input(path)
    if suffix == ".xls":
        raise ValueError("暂不支持 .xls，请另存为 .xlsx 或 .csv 后再导入。")
    raise ValueError("学生名单只支持 .csv 和 .xlsx 文件。")


def normalize_roster_rows(rows: List[Dict[str, object]], headers: List[str]) -> Tuple[List[Dict[str, str]], List[Dict[str, object]]]:
    id_column = find_column(headers, ID_HEADERS)
    name_column = find_column(headers, NAME_HEADERS)
    if not id_column:
        raise ValueError("找不到学生编号列。支持表头：student_id、id、学号、考号、座号等。")
    if not name_column:
        raise ValueError("找不到姓名列。支持表头：name、姓名、学生姓名。")

    roster_rows: List[Dict[str, str]] = []
    validation_rows: List[Dict[str, object]] = []
    seen_ids: Counter = Counter()
    seen_names: Counter = Counter()

    for offset, row in enumerate(rows, start=2):
        raw_values = [str(value or "").strip() for value in row.values()]
        if not any(raw_values):
            validation_rows.append({"severity": "warning", "row_number": offset, "field": "", "message": "空行已跳过"})
            continue
        student_id, id_warnings = normalize_student_id(row.get(id_column))
        name, name_warnings = normalize_name(row.get(name_column))
        if not student_id:
            validation_rows.append({"severity": "error", "row_number": offset, "field": "student_id", "message": "缺少学号"})
        if not name:
            validation_rows.append({"severity": "error", "row_number": offset, "field": "name", "message": "缺少姓名"})
        for message in id_warnings:
            validation_rows.append({"severity": "warning", "row_number": offset, "field": "student_id", "message": message})
        for message in name_warnings:
            validation_rows.append({"severity": "warning", "row_number": offset, "field": "name", "message": message})
        if student_id and name:
            roster_rows.append({"student_id": student_id, "name": name})
            seen_ids[student_id] += 1
            seen_names[name] += 1

    for student_id, count in seen_ids.items():
        if count > 1:
            validation_rows.append({"severity": "error", "row_number": "", "field": "student_id", "message": f"学号重复：{student_id}"})
    for name, count in seen_names.items():
        if count > 1:
            validation_rows.append({"severity": "warning", "row_number": "", "field": "name", "message": f"可能有重名：{name}"})
    duplicate_ids = {student_id for student_id, count in seen_ids.items() if count > 1}
    if duplicate_ids:
        roster_rows = [row for row in roster_rows if row["student_id"] not in duplicate_ids]
    return roster_rows, validation_rows


def update_classes_index(class_name: str, student_count: int, updated_at: str, classes_root: Path = CLASSES_ROOT, class_id: str = "") -> None:
    index = load_classes_index(classes_root)
    classes = list(index.get("classes", []))
    if not class_id:
        try:
            class_id = str(find_class_entry(classes_root, class_name)["class_id"])
        except FileNotFoundError:
            class_id = _next_available_class_id(classes_root, classes)
    _upsert_entry(classes, _entry(class_id, class_name, student_count, updated_at))
    classes.sort(key=lambda item: str(item.get("class_id", "")))
    save_classes_index(classes_root, {"classes": classes})


def import_roster(input_path: Path, class_name: str, notes: str = "", classes_root: Path = CLASSES_ROOT) -> Dict[str, object]:
    class_name = normalize_class_name(class_name)
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"找不到学生名单文件：{input_path}")
    entry = _get_or_create_class_entry(class_name, classes_root)
    class_id = str(entry["class_id"])
    target_dir = classes_root / class_id

    rows, headers = read_roster_input(input_path)
    roster_rows, validation_rows = normalize_roster_rows(rows, headers)
    errors = [row for row in validation_rows if row["severity"] == "error"]
    if errors:
        write_dicts(target_dir / "roster_validation_report.csv", ["severity", "row_number", "field", "message"], validation_rows)
        raise ValueError(f"名单导入失败，共 {len(errors)} 个错误。请查看：{target_dir / 'roster_validation_report.csv'}")

    timestamp = now_iso()
    target_dir.mkdir(parents=True, exist_ok=True)
    write_dicts(target_dir / "roster.csv", ["student_id", "name"], roster_rows)
    write_dicts(target_dir / "roster_validation_report.csv", ["severity", "row_number", "field", "message"], validation_rows)
    _write_class_metadata(
        target_dir,
        class_id,
        class_name,
        len(roster_rows),
        timestamp,
        source_file=input_path.name,
        notes=notes,
    )
    update_classes_index(class_name, len(roster_rows), timestamp, classes_root, class_id=class_id)
    return {"class_id": class_id, "class_name": class_name, "student_count": len(roster_rows), "warnings": len(validation_rows)}


def load_roster(class_name: str, classes_root: Path = CLASSES_ROOT) -> Dict[str, str]:
    entry = find_class_entry(classes_root, class_name)
    roster_path = classes_root / str(entry["class_id"]) / "roster.csv"
    if not roster_path.exists():
        raise FileNotFoundError(f"班级“{entry['class_name']}”还没有学生名单，请先导入 roster。")
    roster: Dict[str, str] = {}
    with roster_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ["student_id", "name"]:
            raise ValueError(f"{roster_path} 的表头必须是：student_id,name")
        for index, row in enumerate(reader, start=2):
            student_id, _warnings = normalize_student_id(row.get("student_id"))
            name, _warnings = normalize_name(row.get("name"))
            if not student_id or not name:
                raise ValueError(f"{roster_path} 第 {index} 行缺少学号或姓名。")
            if student_id in roster:
                raise ValueError(f"{roster_path} 第 {index} 行学号重复：{student_id}")
            roster[student_id] = name
    return roster


def match_student(class_name: str, recognized_student_id: object, classes_root: Path = CLASSES_ROOT) -> Dict[str, object]:
    student_id, warnings = normalize_student_id(recognized_student_id)
    roster = load_roster(class_name, classes_root)
    if student_id in roster:
        return {"matched": True, "student_id": student_id, "name": roster[student_id], "message": ""}
    message = "学生名单中找不到该学号"
    if warnings:
        message = f"{message}；{'；'.join(warnings)}"
    return {"matched": False, "student_id": student_id, "name": "", "message": message}


def validate_class(class_name: str, classes_root: Path = CLASSES_ROOT) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    try:
        target_dir = class_dir(class_name, classes_root)
    except FileNotFoundError as exc:
        return [{"severity": "error", "row_number": "", "field": "class", "message": str(exc)}]
    roster_path = target_dir / "roster.csv"
    if not roster_path.exists():
        return [{"severity": "error", "row_number": "", "field": "roster.csv", "message": "找不到 roster.csv，请先导入学生名单。"}]

    seen_ids: Counter = Counter()
    seen_names: Counter = Counter()
    with roster_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ["student_id", "name"]:
            rows.append({"severity": "error", "row_number": 1, "field": "", "message": "表头必须是 student_id,name"})
            return rows
        for index, row in enumerate(reader, start=2):
            student_id, id_warnings = normalize_student_id(row.get("student_id"))
            name, name_warnings = normalize_name(row.get("name"))
            if not student_id:
                rows.append({"severity": "error", "row_number": index, "field": "student_id", "message": "缺少学号"})
            if not name:
                rows.append({"severity": "error", "row_number": index, "field": "name", "message": "缺少姓名"})
            for message in id_warnings:
                rows.append({"severity": "warning", "row_number": index, "field": "student_id", "message": message})
            for message in name_warnings:
                rows.append({"severity": "warning", "row_number": index, "field": "name", "message": message})
            if student_id:
                seen_ids[student_id] += 1
            if name:
                seen_names[name] += 1
    for student_id, count in seen_ids.items():
        if count > 1:
            rows.append({"severity": "error", "row_number": "", "field": "student_id", "message": f"学号重复：{student_id}"})
    for name, count in seen_names.items():
        if count > 1:
            rows.append({"severity": "warning", "row_number": "", "field": "name", "message": f"可能有重名：{name}"})
    return rows


def list_classes(classes_root: Path = CLASSES_ROOT) -> List[Dict[str, object]]:
    ensure_classes_root(classes_root)
    return list(load_classes_index(classes_root).get("classes", []))


def set_default_class(class_name: str, classes_root: Path = CLASSES_ROOT) -> None:
    entry = find_class_entry(classes_root, class_name)
    load_roster(str(entry["class_name"]), classes_root)
    write_json(
        classes_root / "default_class.json",
        {"default_class_id": entry["class_id"], "default_class_name": entry["class_name"]},
    )


def get_default_class(classes_root: Path = CLASSES_ROOT) -> str:
    ensure_classes_root(classes_root)
    data = read_json(classes_root / "default_class.json", {"default_class_id": "", "default_class_name": ""})
    default_name = str(data.get("default_class_name", "")).strip()
    if default_name:
        return default_name
    default_id = str(data.get("default_class_id", "")).strip()
    if default_id:
        index = load_classes_index(classes_root)
        for item in index.get("classes", []):
            if item.get("class_id") == default_id:
                return str(item.get("class_name", ""))
    return ""


def format_time(value: object) -> str:
    text = str(value or "")
    return text.replace("T", " ")[:16] if text else ""


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="管理班级学生名单。")
    subparsers = parser.add_subparsers(dest="command", required=True)

    import_parser = subparsers.add_parser("import", help="导入 CSV 或 XLSX 学生名单。")
    import_parser.add_argument("--file", required=True, type=Path)
    import_parser.add_argument("--class-name", required=True)
    import_parser.add_argument("--notes", default="")

    subparsers.add_parser("list", help="查看已导入班级。")

    default_parser = subparsers.add_parser("set-default", help="设置默认班级。")
    default_parser.add_argument("--class-name", required=True)

    subparsers.add_parser("default", help="查看默认班级。")

    show_parser = subparsers.add_parser("show", help="显示某个班级的学生名单。")
    show_parser.add_argument("--class-name", required=True)

    validate_parser = subparsers.add_parser("validate", help="检查某个班级学生名单。")
    validate_parser.add_argument("--class-name", required=True)
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    ensure_classes_root()

    if args.command == "import":
        result = import_roster(args.file, args.class_name, args.notes)
        print(f"已导入班级：{result['class_name']}，学生 {result['student_count']} 人。")
        print(f"内部目录：classes/{result['class_id']}/")
        if result["warnings"]:
            print("已生成名单检查报告 roster_validation_report.csv，请查看其中的提醒。")
        return 0

    if args.command == "list":
        classes = list_classes()
        print("已导入班级：")
        if not classes:
            print("（暂无）")
        for item in classes:
            print(f"- {item['class_name']}：{item.get('student_count', 0)} 人，更新时间 {format_time(item.get('updated_at'))}")
        return 0

    if args.command == "set-default":
        set_default_class(args.class_name)
        print(f"默认班级已设置为：{args.class_name}")
        return 0

    if args.command == "default":
        default_class = get_default_class()
        print(f"默认班级：{default_class or '（未设置）'}")
        return 0

    if args.command == "show":
        roster = load_roster(args.class_name)
        print("student_id,name")
        for student_id, name in roster.items():
            print(f"{student_id},{name}")
        return 0

    if args.command == "validate":
        rows = validate_class(args.class_name)
        try:
            report_path = class_dir(args.class_name) / "roster_validation_report.csv"
        except FileNotFoundError:
            report_path = CLASSES_ROOT / "roster_validation_report.csv"
        write_dicts(report_path, ["severity", "row_number", "field", "message"], rows)
        if not rows:
            print("名单检查通过。")
            return 0
        for row in rows:
            print(f"{row['severity']}\trow={row['row_number']}\t{row['field']}\t{row['message']}")
        return 1 if any(row["severity"] == "error" for row in rows) else 0

    parser.error("Unknown command.")
    return 2


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise SystemExit(1)
